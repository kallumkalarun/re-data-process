from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from django.http import FileResponse, HttpResponse
from django_tables2.config import RequestConfig
import django_tables2 as tables
from django_tables2.export.export import TableExport
from django.conf import settings
from django.core.files.storage import FileSystemStorage
import tempfile
import requests
import json
import re
import os



@login_required(login_url="/users/login/")
def gdcsdapiin_view(request):
    if request.method == "POST":
        return render(request, 'tasks/gdc-sd-api-out')
    else:
        return render(request, "tasks/gdc-sd-api-in.html", {'disable_menu' : True})


# ********************

##-------------- After process the file which recieved through browser will send back to browser for user to download srart here -------------##########
@login_required(login_url="/users/login/")
def gdcsdapiout_files(request):
    if request.method == "POST":

        # first_file = request.FILES.get('first_file')  # Get uploaded file
        # second_files = request.FILES.getlist('second_file[]')
        selected_psite = request.POST.get('select_psite') 

        # if not first_file or not second_files:
        #     return render(request, 'tasks/gdc-sd-api-in.html', {"error": "Please select both UPC Excel data  and Tag files."})

        # https://docs.gdc.cancer.gov/API/Users_Guide/Python_Examples/
        # https://docs.gdc.cancer.gov/API/Users_Guide/Search_and_Retrieval/


        # response_data = FilteredQuery()


        response_data = ComplexFilters(selected_psite)

        # if response_data.status_code == 200:
        #     # Parse the JSON response
        #     data = json.dumps(response_data.json(), indent=2)

        data = json.loads(response_data.content)
        # print(data)


        # Call JSON nested structure to list 
        api_list_result = traverse_json(data)
        # print(api_list_result)

        # list of JSON data to row list based on dense ranking
        formatted_data = process_data(api_list_result)
        # print(formatted_data)
        # row = formatted_data[0]
        # print(row[0])
        # print(row[9])

        # Save the formatted data in the session (to retain it across requests)
        request.session['formatted_data'] = formatted_data

    else:

        # Retrieve the saved data for pagination (GET requests)
        formatted_data = request.session.get('formatted_data', [])



    class data_Table(tables.Table):
        rank = tables.columns.Column(verbose_name="Rank")  # Index 0
        submiter_id = tables.columns.Column(verbose_name="Submitter ID")  # Index 1
        disease_type = tables.columns.Column(verbose_name="Disease Type")  # Index 2
        project_id = tables.columns.Column(verbose_name="Project ID")  # Index 3
        primary_site = tables.columns.Column(verbose_name="Primary Site")  # Index 4
        target_project = tables.columns.Column(verbose_name="Target Project")  # Index 5
        sample_type = tables.columns.Column(verbose_name="Sample Type")  # Index 6
        race = tables.columns.Column(verbose_name="Race")  # Index 7
        gender = tables.columns.Column(verbose_name="Gender")  # Index 8
        file_name = tables.columns.Column(verbose_name="File Name")  # Index 9

        class Meta:
            template_name = "django_tables2/bootstrap.html"


    # Convert formatted_data (list of lists) to a list of dictionaries for the table
    formatted_dicts = [
        {
            "rank": row[0] if row[0] is not None else "", 
            "submiter_id": row[1] if row[1] is not None else "", 
            "disease_type": row[2] if row[2] is not None else "", 
            "project_id": row[3] if row[3] is not None else "", 
            "primary_site": row[4] if row[4] is not None else "", 
            "target_project": row[5] if row[5] is not None else "", 
            "sample_type": row[6] if row[6] is not None else "", 
            "race": row[7] if row[7] is not None else "", 
            "gender": row[8] if row[8] is not None else "", 
            "file_name": row[9] if row[9] is not None else "", 
        }
        for row in formatted_data
    ]



    table = data_Table(formatted_dicts)
    # print(type(data_Table))
    # print(data_Table)
    request_config = RequestConfig(request, paginate={"per_page":25})
    request_config.configure(table)

    # Handle the case where there's no data
    if not table.data:
        no_data_message = "No data available for the report."
    else:
        no_data_message = None  # Clear the message if data is present

    context = {
        "table": table,
        "no_data_message": no_data_message,
        'disable_menu': True,
    }
    return render(request, 'tasks/gdc-sd-api-out.html', context)
    
    #   return render(request, 'tasks/gdc-sd-api-out.html', context)
    # else:
    #     
    #     return render(request, "tasks/gdc-sd-api-out.html", {'disable_menu' : True})

  


# creating excel for download funtion, but its not implemented 
def create_excel_workbook(data):
    
    workbook = Workbook()
    sheet = workbook.active  # Get the active sheet

    column_headers = ["Rank", "Submitter ID", "Disease Type", "Project ID", "Primary Site", "Target Project", "Sample Type", "Race", "Gender", "File Name"]  
    for col_idx, header in enumerate(column_headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    return workbook


#**********************

# list of JSON data to row list based on dense ranking
def process_data(api_list_result):
    # Group values by dense ranking
    grouped_data = {}
    for row in api_list_result:
        rank = row[0]
        value = row[1]

        # print(rank)
        # print(value)

        if rank not in grouped_data:
            grouped_data[rank] = []
        grouped_data[rank].append(value)
    
    # Transform grouped data into a list of rows
    # formatted_data = [[rank, ", ".join(values)] for rank, values in grouped_data.items()]
    formatted_data = [[rank] + values for rank, values in grouped_data.items()]
    return formatted_data






# Call JSON nested structure to list 
def traverse_json(data, parent_key='', list_result = []):
    
    if isinstance(data, dict):
        for key, value in data.items():
            current_key = f"{parent_key}.{key}" if parent_key else key
            if isinstance(value, (dict, list)):
                # If the value is a dictionary or list, print the key and recursively call for nested data
                # print(f"{current_key}:")
                # traverse_json(value, current_key)
                traverse_json(value, current_key)

            else:
                # If it's a value, print the key and value pair
                # print(f"{current_key}: {value}")
                # print(f"{key}: {value}")
                # print(f"{value}")

                

                match = re.search(r"\[(\d+)\]", current_key)
                if match:
                    number = match.group(1)
                    # print(f"{number}: {value}")
                    list_result.append([number, value])
                    

    
    elif isinstance(data, list):
        for index, item in enumerate(data):
            current_key = f"{parent_key}[{index}]"
            # traverse_json(item, current_key)
            traverse_json(item, current_key)
            
    return list_result





def ComplexFilters(selected_psite):
    fields = [
    "file_name",
    "cases.submitter_id",
    "cases.project.project_id",
    "cases.samples.sample_type",
    "cases.disease_type",
    "cases.project.primary_site",
    "cases.demographic.race",
    "cases.demographic.gender"
    ]

    fields = ",".join(fields)

    # pSite = ["Lung", "Kidney", "Brain", "Blood"]
    pSite = selected_psite


    files_endpt = "https://api.gdc.cancer.gov/files"

    # This set of filters is nested under an 'and' operator.
    filters = {
        "op": "and",
        "content":[
            {
            "op": "in",
            "content":{
                "field": "cases.project.primary_site",
                # "value": ["Lung", "Kidney", "Brain", "Blood"]
                "value": pSite
                }
            },
            {
            "op": "in",
            "content":{
                "field": "files.experimental_strategy",
                "value": ["RNA-Seq"]
                }
            },
            # {
            # "op": "in",
            # "content":{
            #     "field": "files.data_format",
            #     "value": ["BAM"]
            #     }
            # },
            # {
            # "op": "in",
            # "content":{
            #     "field": "cases.demographic.race",
            #     "value": ["white"]
            #     }
            # },
            # {
            # "op": "in",
            # "content":{
            #     "field": "cases.demographic.gender",
            #     "value": ["female"]
            #     }
            # },
            # {
            # "op": "in",
            # "content":{
            #     "field": "files.analysis.workflow_type",
            #     "value": ["HTSeq - FPKM"]
            #     }
            # }
        ]
    }

    # A POST is used, so the filter parameters can be passed directly as a Dict object.
    params = {
        "filters": filters,
        "fields": fields,
        "format": "JSON",
        "size": "2000",
        "pretty" : "True",
        "sort" : "submitter_id:asc"
        }

    # The parameters are passed to 'json' rather than 'params' in this case
    response = requests.post(files_endpt, headers = {"Content-Type": "application/json"}, json = params)

    return response

# ********************



# ********************

def FilteredQuery():
    fields = [
    "submitter_id",
    "case_id",
    "primary_site",
    "disease_type",
    "diagnoses.vital_status"
    ]

    fields = ",".join(fields)

    cases_endpt = "https://api.gdc.cancer.gov/cases"

    filters = {
        "op": "in",
        "content":{
            "field": "primary_site",
            "value": ["Kidney"]
            }
        }

    # With a GET request, the filters parameter needs to be converted
    # from a dictionary to JSON-formatted string

    params = {
        "filters": json.dumps(filters),
        "fields": fields,
        "format": "JSON",
        "size": "100"
        }

    response = requests.get(cases_endpt, params = params)

    return response



# ********************




def FilesBasedOnFilter():

    files_endpt = "https://api.gdc.cancer.gov/files"

    filters = {
        "op": "and",
        "content":[
            {
            "op": "in",
            "content":{
                "field": "cases.project.primary_site",
                "value": ["Lung"]
                }
            },
            # {
            # "op": "in",
            # "content":{
            #     "field": "cases.demographic.race",
            #     "value": ["white"]
            #     }
            # },
            # {
            # "op": "in",
            # "content":{
            #     "field": "cases.demographic.gender",
            #     "value": ["female"]
            #     }
            # },
            # {
            # "op": "in",
            # "content":{
            #     "field": "files.analysis.workflow_type",
            #     "value": ["HTSeq - FPKM"]
            #     }
            # }
        ]
    }

    # Here a GET is used, so the filter parameters should be passed as a JSON string.

    params = {
        "filters": json.dumps(filters),
        "fields": "file_id",
        "format": "JSON",
        "size": "1000"
        }

    response = requests.get(files_endpt, params = params)


    return response

# ********************


